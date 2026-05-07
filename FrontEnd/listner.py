import os
import traceback
from rediscon import redis_conn
from db import getdb
from openpyxl import load_workbook

# ==========================================
# STANDALONE REDIS LISTENER SERVICE
# ==========================================
# Run as a service on Railway: python listner.py
# Handles job completion events from Redis

print("Starting background Redis listener service...")
pubsub = redis_conn.pubsub()
pubsub.subscribe("batch_complete")

for message in pubsub.listen():
    if message["type"] != "message":
        continue

    # Decode data just in case Redis returns bytes
    raw_data = message["data"]
    request_id = raw_data.decode("utf-8") if isinstance(raw_data, bytes) else raw_data
    print(f"Listener triggered for request_id: {request_id}")

    con = getdb()
    cursor = con.cursor()

    try:
        # Get the data of the jobs that just got completed 
        cursor.execute("""
            SELECT 
                j.id AS job_id,
                j.row_number,
                j.created_at,
                j.job_type,
                jr.output,
                jr.success,
                jr.error
            FROM jobs j
            LEFT JOIN job_results jr
            ON j.id = jr.job_id
            WHERE j.request_id = %s::uuid
            ORDER BY j.created_at ASC;
        """, (request_id,))
        
        job_results = cursor.fetchall()
        
        # Get the file to update
        cursor.execute("""
            SELECT 
                u.filename,
                u.s3_key,
                u.size_bytes,
                u.created_at
            FROM requests r
            JOIN uploads u 
            ON r.upload_id = u.id
            WHERE r.id = %s::uuid;
        """, (request_id,))
        file_info = cursor.fetchone()
        
        # Updating the original uploaded file
        if file_info and os.path.exists(file_info['s3_key']):
            wb = load_workbook(file_info['s3_key'])
            ws = wb.active
            
            for job in job_results:
                row_number = job['row_number']
                
                if int(job['job_type']) == 1:
                    if job['success']:
                        ws.cell(row=row_number, column=3).value = job['output']
                    else:
                        ws.cell(row=row_number, column=3).value = f"Error: {job['error']}"      
                elif int(job['job_type']) == 2:
                    print("job type 2")
                elif int(job['job_type']) == 3:
                    print("job type 3")

            # Update the job status
            cursor.execute("""
                UPDATE requests
                SET status = 'completed'
                WHERE id = %s::uuid
            """, (request_id,))
            
            con.commit()
            wb.save(file_info['s3_key'])
            wb.close()
            print(f"Successfully processed and updated file for request {request_id}")

    except Exception as e:
        print(f"Error in Redis listener processing: {e}")
        traceback.print_exc()
        con.rollback()
    finally:
        cursor.close()
        con.close()