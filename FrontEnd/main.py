import os
import shutil
import traceback
from typing import Optional
import requests
from datetime import datetime
import json
import google.generativeai as genai
import sqlparse
import re
import threading
from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook
from fastapi.staticfiles import StaticFiles
from contextlib import asynccontextmanager
# Your internal imports
import data
import util
import auth
from db import getdb
from rediscon import redis_conn
from models.models import Job, user, userLogin

# URL for job queuing api 
api2url = data.API2URL

# Replace with FireBase cloud Storage 
UPLOAD_DIRECTORY = os.getenv("STORAGE_PATH", "local_storage")

os.makedirs(UPLOAD_DIRECTORY, exist_ok=True)
ALLOWED_TABLES = {"requests", "jobs", "job_results"}


rail = os.getenv("ENV")
if rail is not None and rail == "RAILWAY":
    genai.configure(api_key= os.getenv("GEMINI_KEY_RAIL"))
else:
    from local_config import GEMINI_KEY_LOCAL
    genai.configure(api_key= GEMINI_KEY_LOCAL)


model = genai.GenerativeModel("gemini-2.5-flash")

# ==========================================
# FASTAPI INITIALIZATION
# ==========================================
# Note: Redis listener is now a separate service (listner.py)
# Run it independently: python listner.py

def run_redis_listener():
    print("Starting background Redis listener...")
    pubsub = redis_conn.pubsub()
    pubsub.subscribe("batch_complete")

    for message in pubsub.listen():
        if message["type"] != "message":
            continue

        # Decode data just in case Redis returns bytes
        raw_data = message["data"]
        request_id = raw_data.decode("utf-8") if isinstance(raw_data, bytes) else raw_data
        print(f"Listener triggered for request_id: {request_id}")

        con = getdb()
        cursor = con.cursor()

        try:
            # Get the data of the jobs that just got completed 
            cursor.execute("""
                SELECT 
                    j.id AS job_id,
                    j.row_number,
                    j.created_at,
                    j.job_type,
                    jr.output,
                    jr.success,
                    jr.error
                FROM jobs j
                LEFT JOIN job_results jr
                ON j.id = jr.job_id
                WHERE j.request_id = %s::uuid
                ORDER BY j.created_at ASC;
            """, (request_id,))
            
            job_results = cursor.fetchall()
            
            # Get the file to update
            cursor.execute("""
                SELECT 
                    u.filename,
                    u.s3_key,
                    u.size_bytes,
                    u.created_at
                FROM requests r
                JOIN uploads u 
                ON r.upload_id = u.id
                WHERE r.id = %s::uuid;
            """, (request_id,))
            file_info = cursor.fetchone()
            
            # Updating the original uploaded file
            if file_info and os.path.exists(file_info['s3_key']):
                wb = load_workbook(file_info['s3_key'])
                ws = wb.active
                
                for job in job_results:
                    row_number = job['row_number']
                    
                    if int(job['job_type']) == 1:
                        if job['success']:
                            ws.cell(row=row_number, column=3).value = job['output']
                        else:
                            ws.cell(row=row_number, column=3).value = f"Error: {job['error']}"      
                    elif int(job['job_type']) == 2:
                        print("job type 2")
                    elif int(job['job_type']) == 3:
                        print("job type 3")

                # Update the job status
                cursor.execute("""
                    UPDATE requests
                    SET status = 'completed'
                    WHERE id = %s::uuid
                """, (request_id,))
                
                con.commit()
                wb.save(file_info['s3_key'])
                wb.close()
                print(f"Successfully processed and updated file for request {request_id}")

        except Exception as e:
            print(f"Error in Redis listener processing: {e}")
            con.rollback()
        finally:
            cursor.close()
            con.close()

# ==========================================
# 2. FASTAPI LIFESPAN
# ==========================================
@asynccontextmanager
async def lifespan(app: FastAPI):
    # Fire up the listener in a daemon thread so it doesn't block the API
    # Daemon=True means it will automatically die when the FastAPI server stops
    listener_thread = threading.Thread(target=run_redis_listener, daemon=True)
    listener_thread.start()
    yield
    # Cleanup logic (if any) would go here when the server shuts down

# Initialize FastAPI with the lifespan
app = FastAPI(lifespan=lifespan)

#app = FastAPI()
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")


# ==========================================
# API ENDPOINTS
# ==========================================

@app.get("/")
async def home(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

@app.post("/upload")
async def handele_upload(
    current_user_id = Depends(auth.get_current_user),
    user_file: UploadFile = File(...),
    option1: bool = Form(False),
    option2: bool = Form(False),
    option3: bool = Form(False),
    option4: bool = Form(False),
    
    special_job: bool = Form(False), 
    special_date: Optional[str] = Form(None),
    special_time: Optional[str] = Form(None),
    retry_number: int = Form(0)
    
):
    print(f"\n=== UPLOAD ENDPOINT CALLED ===")
    con = getdb()
    cursor = con.cursor()
    
    #speical job validation 
    if special_job:
        missing_fields = []
        if not special_date:
            missing_fields.append("special_date")
        if not special_time:
            missing_fields.append("special_time")
        if retry_number < 0:
            raise HTTPException(status_code=400, detail="retry_number must be a non-negative integer.")
        
        if missing_fields:
            raise HTTPException(status_code=400, detail=f"Missing required fields for special job: {', '.join(missing_fields)}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = os.path.basename(user_file.filename) 
    unique_filename = f"{current_user_id}_{timestamp}_{safe_name}"
    file_location = os.path.join(UPLOAD_DIRECTORY, unique_filename)
    
    try:
        with open(file_location, "wb+") as file_object:
            shutil.copyfileobj(user_file.file, file_object)
            
        file_size_bytes = os.path.getsize(file_location)

        insert_query = """
            INSERT INTO uploads (user_id, filename, s3_key, size_bytes)
            VALUES (%s::uuid, %s, %s, %s)
            RETURNING id;
        """
        cursor.execute(insert_query, (current_user_id, unique_filename, file_location, file_size_bytes))
        new_upload_id = cursor.fetchone()['id']
        con.commit()
        
        request_type = await get_combination_id(option1, option2, option3)
        rqest_id = await createRequest(current_user_id, new_upload_id, file_location, request_type , special_job, special_date, special_time, retry_number)
        
        return {
                "message": "Upload started",
                "request_id": str(rqest_id)
            }
        
    except Exception as e:
        con.rollback()
        traceback.print_exc() 
        raise HTTPException(status_code=500, detail="Failed to upload and save to database.")
    finally:
        cursor.close()
        con.close()

async def createRequest(current_user_id: str, upload_id: str, current_path: str, request_type: str , special_job: bool, special_date: Optional[str], special_time: Optional[str], retry_number: int = 0):
    con = getdb()
    cursor = con.cursor()
    try: 
        wb = load_workbook(current_path)
        ws = wb.active
        total_rows = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True))
        
        insert_query = """INSERT INTO requests (user_id, upload_id, name, total_jobs , isschedule)
                        VALUES (%s::uuid, %s::uuid, %s, %s, %s)
                        RETURNING id;"""
        
        cursor.execute(insert_query, (current_user_id, upload_id, os.path.basename(current_path), total_rows, special_job))
        new_rquest_id = cursor.fetchone()['id']
        con.commit()
        
        redis_conn.set(f"batch:{new_rquest_id}:remaining", total_rows)
        await createJobs(current_user_id, new_rquest_id, current_path, request_type , special_job, special_date, special_time, retry_number)
        
        return new_rquest_id
        
    except Exception as e:
        con.rollback()
        traceback.print_exc() 
        raise HTTPException(status_code=500, detail="Failed to upload and save to database.")
    finally:
        cursor.close()
        con.close()

async def createJobs(current_user_id: str, request_id: str, current_path: str, job_type: str ,special_job: bool = False, special_date: Optional[str] = None, special_time: Optional[str] = None, retry_number: int = 0):
    con = getdb()
    cursor = con.cursor()
    try:
        wb = load_workbook(current_path)
        ws = wb.active
        
        count = 2 
        for row in ws.iter_rows(min_row=2, values_only=True):
            pan_id = row[0]
            pass_id = row[1]
            
            insert_query = """INSERT INTO jobs (request_id, row_number ,job_type) 
                            VALUES (%s::uuid, %s , %s)
                            RETURNING id; """
            cursor.execute(insert_query, (request_id, count, job_type))
            new_job_id = cursor.fetchone()['id']
            
            paylod = {
                "job_id": new_job_id,
                "pan_id": pan_id,
                "pass_id": pass_id,
                "job_code": job_type,
                "request_id": request_id,
                "schedule_date": special_date,
                "schedule_time": special_time,
                "retry_number": retry_number,
                "special_job": special_job
            }
            
            requests.post(api2url, json=paylod)
            count += 1
            
        con.commit()
        
    except Exception as e:
        con.rollback()
        traceback.print_exc() 
        raise HTTPException(status_code=500, detail="Failed to upload and save to database.")
    finally:
        cursor.close()
        con.close()
        wb.close()
        
@app.get("/requests/{request_id}/status")
async def request_status(request_id: str):
    con = getdb()
    cursor = con.cursor()
    try:
        cursor.execute("""
            SELECT id, name, total_jobs, status
            FROM requests
            WHERE id = %s::uuid
        """, (request_id,))

        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Request not found")

        remaining = redis_conn.get(f"batch:{request_id}:remaining")
        remaining = 0 if remaining is None else int(remaining)
        proceessed_job = row["total_jobs"] - remaining

        return {
            "request_id": str(row["id"]),
            "name": row["name"],
            "status": row["status"],
            "total_jobs": row["total_jobs"],
            "completed_jobs": proceessed_job
        }
    finally:
        cursor.close()
        con.close()
        
@app.get("/requests/recent")
async def get_recent_requests(current_user_id = Depends(auth.get_current_user)):
    con = getdb()
    cursor = con.cursor()
    try:
        cursor.execute("""
            SELECT id, name, total_jobs, status, created_at
            FROM requests
            WHERE user_id = %s::uuid
            AND created_at >= NOW() - INTERVAL '2 days'
            ORDER BY created_at ASC
        """, (current_user_id,))

        rows = cursor.fetchall()
        results = []
        for r in rows:
            results.append({
                "id": str(r["id"]),
                "name": r["name"],
                "status": r["status"],
                "total_jobs": r["total_jobs"]
            })
        return results
    finally:
        cursor.close()
        con.close()    
        
@app.get("/requests/{request_id}/download")
async def download_request_file(request_id: str, current_user_id = Depends(auth.get_current_user)):
    con = getdb()
    cursor = con.cursor()
    try:
        cursor.execute("""
            SELECT u.filename, u.s3_key
            FROM requests r
            JOIN uploads u ON r.upload_id = u.id
            WHERE r.id = %s::uuid
            AND r.user_id = %s::uuid
        """, (request_id, current_user_id))

        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="File not found")

        return FileResponse(
            path=row["s3_key"],
            filename=row["filename"],
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    finally:
        cursor.close()
        con.close()
        
@app.post("/userreg")
async def user_registration(user: user):
    con = getdb()
    cursor = con.cursor()
    try:
        password_bytes = len(user.password.encode("utf-8"))
        if password_bytes > 72:
            raise HTTPException(status_code=400, detail="Password too long (max 72 bytes)")
        
        hashed_password = util.hash(user.password)
        
        cursor.execute("""
            INSERT INTO users (email, hashed_password, name)
            VALUES (%s, %s, %s)
            RETURNING id;
        """, (user.email, hashed_password, user.name))

        new_user_id = cursor.fetchone()['id']
        con.commit()
        return {"message": "User registered successfully", "user_id": str(new_user_id)}
    except Exception as e:
        con.rollback()
        traceback.print_exc() 
        raise HTTPException(status_code=500, detail="Failed to register user, maybe email already exists.")
    finally:
        cursor.close()
        con.close()
        
@app.get("/users/me")        
async def get_user(user_id = Depends(auth.get_current_user)):
    con = getdb()
    cursor = con.cursor()
    try: 
        cursor.execute("""
            SELECT id, email, name
            FROM users
            WHERE id = %s::uuid
        """, (user_id,))

        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="User not found")

        return {
            "id": str(row["id"]),
            "email": row["email"],
            "name": row["name"]
        }
    finally:
        cursor.close()
        con.close()
        
@app.post('/login')
def login(user_cred: OAuth2PasswordRequestForm = Depends()):
    con = getdb()
    cursor = con.cursor()
    try:
        cursor.execute("""
            SELECT id, email, hashed_password
            FROM users
            WHERE email = %s
        """, (user_cred.username,))

        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="User not found")

        password_valid = util.verify(user_cred.password, row["hashed_password"])
        if not password_valid:
            raise HTTPException(status_code=401, detail="Invalid credentials")

        access_token = auth.create_acess_token(data={"user_id": str(row["id"])})
        return {"access_token": access_token, "token_type": "bearer"}
    finally:
        cursor.close()
        con.close()


# ==========================================
# 4. SCHEDULED & ADDITIONAL ENDPOINTS
# ==========================================

@app.get("/requests/scheduled")
async def get_scheduled_requests(current_user_id = Depends(auth.get_current_user)):
    """
    Get all scheduled requests from the last 5 days for the current user
    """
    con = getdb()
    cursor = con.cursor()
    try:
        cursor.execute("""
            SELECT id, name, total_jobs, status, created_at, isschedule
            FROM requests
            WHERE user_id = %s::uuid
            AND isschedule = true
            AND created_at >= NOW() - INTERVAL '5 days'
            ORDER BY created_at DESC
        """, (current_user_id,))

        rows = cursor.fetchall()
        results = []
        for r in rows:
            remaining = redis_conn.get(f"batch:{r['id']}:remaining")
            remaining = 0 if remaining is None else int(remaining)
            completed = r["total_jobs"] - remaining
            
            results.append({
                "id": str(r["id"]),
                "name": r["name"],
                "status": r["status"],
                "total_jobs": r["total_jobs"],
                "completed_jobs": completed,
                "created_at": r["created_at"].isoformat() if r["created_at"] else None,
                "is_scheduled": r["isschedule"]
            })
        return results
    finally:
        cursor.close()
        con.close()


@app.get("/requests/{request_id}/results")
async def get_request_results(request_id: str, current_user_id = Depends(auth.get_current_user)):
    """
    Get summary of results for a specific request (completed jobs with outputs)
    """
    con = getdb()
    cursor = con.cursor()
    try:
        # Verify user owns this request
        cursor.execute("""
            SELECT id, name, total_jobs, status FROM requests WHERE id = %s::uuid AND user_id = %s::uuid
        """, (request_id, current_user_id))
        
        req = cursor.fetchone()
        if not req:
            raise HTTPException(status_code=403, detail="Unauthorized access to request")
        
        cursor.execute("""
            SELECT 
                j.row_number,
                j.job_type,
                jr.output,
                jr.success,
                jr.error
            FROM jobs j
            LEFT JOIN job_results jr ON j.id = jr.job_id
            WHERE j.request_id = %s::uuid
            AND jr.job_id IS NOT NULL
            ORDER BY j.row_number ASC
        """, (request_id,))

        rows = cursor.fetchall()
        successful = sum(1 for r in rows if r["success"])
        failed = sum(1 for r in rows if not r["success"])
        
        results = []
        for r in rows:
            results.append({
                "row_number": r["row_number"],
                "job_type": r["job_type"],
                "output": r["output"],
                "success": r["success"],
                "error": r["error"]
            })
        
        return {
            "request_id": str(req["id"]),
            "name": req["name"],
            "total_jobs": req["total_jobs"],
            "successful_jobs": successful,
            "failed_jobs": failed,
            "status": req["status"],
            "results": results
        }
    finally:
        cursor.close()
        con.close()


@app.get("/stats")
async def get_user_stats(current_user_id = Depends(auth.get_current_user)):
    """
    Get comprehensive statistics for the current user's requests
    """
    con = getdb()
    cursor = con.cursor()
    try:
        cursor.execute("""
            SELECT 
                COUNT(*) as total_requests,
                SUM(CASE WHEN status = 'completed' THEN 1 ELSE 0 END) as completed_requests,
                SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END) as pending_requests,
                SUM(CASE WHEN status = 'failed' THEN 1 ELSE 0 END) as failed_requests,
                SUM(CASE WHEN isschedule = true THEN 1 ELSE 0 END) as scheduled_requests,
                SUM(total_jobs) as total_jobs_processed
            FROM requests
            WHERE user_id = %s::uuid
        """, (current_user_id,))
        
        stats = cursor.fetchone()
        
        cursor.execute("""
            SELECT SUM(size_bytes) as total_storage_used
            FROM uploads
            WHERE user_id = %s::uuid
        """, (current_user_id,))
        
        storage = cursor.fetchone()
        
        return {
            "total_requests": stats["total_requests"] or 0,
            "completed_requests": stats["completed_requests"] or 0,
            "pending_requests": stats["pending_requests"] or 0,
            "failed_requests": stats["failed_requests"] or 0,
            "scheduled_requests": stats["scheduled_requests"] or 0,
            "total_jobs_processed": stats["total_jobs_processed"] or 0,
            "total_storage_used_bytes": storage["total_storage_used"] or 0
        }
    finally:
        cursor.close()
        con.close()

async def get_combination_id(opt1, opt2, opt3):
    total = 0
    if opt1: total += 1
    if opt2: total += 2
    if opt3: total += 4
    return total




##ai Implemteiodn 

def validate_sql_safety(sql: str):
    parsed = sqlparse.parse(sql)

    if len(parsed) != 1:
        raise HTTPException(status_code=400, detail="Multiple SQL statements not allowed")

    stmt = parsed[0]

    if stmt.get_type() != "SELECT":
        raise HTTPException(status_code=400, detail="Only SELECT allowed")

    sql_lower = sql.lower()

    # Basic injection protection
    forbidden = [";", "--", "/*", "*/", "drop", "delete", "insert", "update", "alter"]
    if any(f in sql_lower for f in forbidden):
        raise HTTPException(status_code=400, detail="Unsafe SQL detected")

    # Check allowed tables
    if not any(tbl in sql_lower for tbl in ALLOWED_TABLES):
        raise HTTPException(status_code=400, detail="Querying unknown tables not allowed")
        
    return sql

##def proumpt
def build_prompt(user_query: str, user_id: str):
    return f"""
You are an AI assistant for a job processing system.

Database schema:

requests(id, user_id, name, total_jobs, status, created_at, isschedule)
jobs(id, request_id, row_number, job_type, status, created_at)
job_results(job_id, output, success, error)

STRICT RULES:
- ONLY generate a single SELECT query
- NO semicolons
- NO joins unless necessary
- ALWAYS use proper table names (requests, jobs, job_results)
- DO NOT include user_id in query (backend will handle it)
- Keep query simple

Supported actions:
- retry_failed_jobs

Return ONLY valid JSON. No explanation, no text, no markdown.

Examples:

Query:
{{
  "type": "query",
  "sql": "SELECT * FROM requests"
}}

Action:
{{
  "type": "action",
  "action": "retry_failed_jobs",
  "params": {{"request_id": "uuid"}}
}}

User query:
"{user_query}"
"""

##safe check
def is_safe_sql(sql: str):
    sql = sql.lower().strip()
    return sql.startswith("select")

@app.post("/ai/query")
async def ai_query(request: Request, current_user_id=Depends(auth.get_current_user)):
    body = await request.json()
    user_query = body.get("query")

    if not user_query:
        raise HTTPException(status_code=400, detail="Query is required")

    prompt = build_prompt(user_query, current_user_id)

    try:
        response = model.generate_content(
            prompt,
            generation_config={
                "temperature": 0
            }
        )

        content = response.text

        # Parse JSON safely
        

        try:
            json_match = re.search(r"\{.*\}", content, re.DOTALL)
            if not json_match:
                raise ValueError("No JSON found")

            ai_output = json.loads(json_match.group())
        except:
            raise HTTPException(status_code=500, detail=f"AI returned invalid JSON: {content}")

        # QUERY MODE
        # ========================
        if ai_output["type"] == "query":
            raw_sql = ai_output.get("sql")

            if not raw_sql:
                raise HTTPException(status_code=400, detail="No SQL provided")

            # 1. Validate the syntax and safety
            validate_sql_safety(raw_sql)

            # 2. Build the CTE Sandbox
            # We redefine the tables as temporary, filtered views for this execution only.
            secure_query = f"""
            WITH requests AS (
                SELECT * FROM public.requests WHERE user_id = %s::uuid
            ),
            jobs AS (
                SELECT j.* FROM public.jobs j
                JOIN public.requests r ON j.request_id = r.id
                WHERE r.user_id = %s::uuid
            ),
            job_results AS (
                SELECT jr.* FROM public.job_results jr
                JOIN public.jobs j ON jr.job_id = j.id
                JOIN public.requests r ON j.request_id = r.id
                WHERE r.user_id = %s::uuid
            )
            {raw_sql} 
            """

            con = getdb()
            cursor = con.cursor()

            try:
                # 3. Execute securely. We pass the user_id 3 times for the 3 CTE filters.
                cursor.execute(secure_query, (current_user_id, current_user_id, current_user_id))
                
                rows = cursor.fetchall()
                # If your cursor doesn't return dicts natively, you might need:
                # columns = [col[0] for col in cursor.description]
                # result = [dict(zip(columns, row)) for row in rows]
                result = [dict(row) for row in rows] 

                return {
                    "type": "query_result",
                    "data": result,
                    "executed_sql": raw_sql  # Keep this clean for the frontend
                }

            except Exception as db_e:
                # This will catch any remaining database-level syntax errors
                print(f"Database Execution Error: {db_e}")
                raise HTTPException(status_code=400, detail="AI generated an invalid SQL query.")
            finally:
                cursor.close()
                con.close()

        # ========================
        # ACTION MODE
        # ========================
        elif ai_output["type"] == "action":
            action = ai_output.get("action")
            params = ai_output.get("params", {})

            # Example: retry failed jobs
            if action == "retry_failed_jobs":
                request_id = params.get("request_id")

                if not request_id:
                    raise HTTPException(status_code=400, detail="request_id required")

                con = getdb()
                cursor = con.cursor()

                try:
                    cursor.execute("""
                        SELECT j.id
                        FROM jobs j
                        LEFT JOIN job_results jr ON j.id = jr.job_id
                        WHERE j.request_id = %s::uuid
                        AND (jr.success = false OR jr.success IS NULL)
                    """, (request_id,))

                    jobs = cursor.fetchall()

                    for job in jobs:
                        requests.post(api2url, json={
                            "job_id": job["id"],
                            "request_id": request_id,
                            "retry": True,
                            "special_job": False
                        })

                    return {
                        "type": "action_result",
                        "message": f"Retried {len(jobs)} jobs"
                    }

                finally:
                    cursor.close()
                    con.close()

            return {"type": "action_result", "message": "Unknown action"}

        else:
            raise HTTPException(status_code=400, detail="Invalid AI response")

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="AI processing failed")